from openpyxl import Workbook
from openpyxl.styles import Font


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def export_to_excel(
        prices: list[dict],
        variants: list[dict],
        path: str = "prices.xlsx",
) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Цены"
    ws.append(["Предмет", "Мин. цена", "Медиана", "Продаж за сутки"])

    for row in prices:
        ws.append([row["name"], row["lowest"], row["median"], row["volume"]])

    for r in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in r:
            cell.number_format = "#,##0.00"

    for r in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in r:
            cell.number_format = "#,##0"

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    _style_header(ws)

    ws2 = wb.create_sheet("Состояния")
    ws2.append(["Исходный предмет", "Вариант", "Состояние", "Мин. цена"])

    for row in variants:
        ws2.append([row["source"], row["variant"], row["exterior"], row["min_price"]])

    for r in ws2.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in r:
            cell.number_format = "#,##0.00"

    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 32
    ws2.column_dimensions["D"].width = 14
    _style_header(ws2)

    wb.save(path)
    print(f"\nФайл сохранён: {path}")